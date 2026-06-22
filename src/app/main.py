"""
FastAPI Gateway - Lojistik Optimizasyon Sistemi
Person D: Sistem Mimari ve Arayuz Gelistiricisi

Moduller:
- /api/optimize        -> OR-Tools rota optimizasyonu (SYNC - deprecated, BackgroundTasks kullan)
- /api/optimize/async  -> Arka plan async optimizasyon (tercih edilen)
- /api/predict         -> LSTM talep tahmini (placeholder)
- /api/fleet           -> Filo atama durumu
- /api/tm-status       -> Transfer merkezi kapasite izleme
- /api/excel           -> Excel cikti uretimi
- /api/demand          -> Sayfali talep verisi
"""

from __future__ import annotations

import datetime
import json
import logging
import math
import os
import sys
import threading
from functools import lru_cache
from pathlib import Path
from typing import Annotated, Optional

from fastapi import BackgroundTasks, Depends, FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware  # FIX #38
from fastapi.responses import ORJSONResponse  # FIX #37
from pydantic import BaseModel, Field, field_validator
from pydantic import ConfigDict

logger = logging.getLogger(__name__)

# Proje kok dizinini sys.path'e ekle (container icinde src/ dogrudan erisilebilir)
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
_SRC_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)
if _SRC_DIR not in sys.path:
    sys.path.insert(0, _SRC_DIR)

from main import run_pipeline, result_to_dict
from utils.data_loader import load_input as _load_input_raw, available_dates, DataContractError
from app.job_manager import create_job, set_running, set_completed, set_failed, get_job as _get_job, get_job_for_date as _get_job_for_date

# FIX #10 - Eszamanli is sayisi sinirlama (semaphore).
# NOT: Worker'lar FastAPI BackgroundTasks ile calistirilir; ayri bir
# ThreadPoolExecutor kullanilmaz (onceki olu _executor kaldirildi).
_MAX_CONCURRENT_JOBS = 4
_semaphore = threading.Semaphore(_MAX_CONCURRENT_JOBS)

app = FastAPI(
    title="Lojistik Optimizasyon API",
    description="Teknofest LogiAI - Karar Destek Sistemi Backend",
    version="1.0.0",
    default_response_class=ORJSONResponse,  # FIX #37
)

# FIX #38 - GZip sikistirma
app.add_middleware(GZipMiddleware, minimum_size=1000)

# FIX #5 - CORS allow_origins env var'dan (production'da kisitli)
_allowed_origins = os.environ.get(
    "ALLOWED_ORIGINS", "http://localhost:8501"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Proje kok dizini: src/app/main.py -> ../../ (proje koku)
# Kaptan yapisi: girdi JSON (Kisi A ciktisi) artik data/processed altinda;
# data/raw yalnizca ham giris Excel'lerini barindirir.
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(_PROJECT_ROOT, "data", "processed"))
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", os.path.join(_PROJECT_ROOT, "data", "processed"))
INPUT_JSON = os.environ.get("INPUT_JSON", os.path.join(DATA_DIR, "logiai_mvp_input.json"))

# FIX #34 - TM kapasite/asim/ceza heuristikleri kaldirildi.
# Sartname (MVP/Dataset A): "Transfer merkezi kisiti yoktur" ve "sinirsiz
# alaniniz oldugunu varsayabilirsiniz." Optimizasyon motoru TM kapasitesini
# kullanmiyor; bu yuzden /api/tm-status artik yalnizca gercek daily_demand'den
# turetilen sehir bazli talep yogunlugunu (giren/cikan desi) doner.

# --------------------------------------------------
#  FIX #9 & #39 - load_input mtime-duyarli cache
# --------------------------------------------------
# FIX #39 - Dosya mtime'i cache anahtarinin parcasidir; data/raw volume mount
# edildiginden JSON guncellenince mtime degisir, yeni anahtar = cache miss =
# otomatik yeniden okuma. API restart gerekmez. Ayni mtime anahtari tum tureyen
# cache'lerde (_build_demand_index, _load_city_coords, _fleet_skeleton, _tm_density)
# kullanildigindan girdi degisince hepsi birlikte tazelenir (FIX #41 ile uyumlu).

def _path_mtime(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except OSError:
        return 0.0


def _input_mtime() -> float:
    """INPUT_JSON'un guncel mtime'i - tureyen cache'lerin tazeleme anahtari."""
    return _path_mtime(INPUT_JSON)


@lru_cache(maxsize=2)
def _load_input_cached(path: str, mtime: float) -> dict:
    return _load_input_raw(path)


def load_input(path: str) -> dict:
    """JSON'u diskten okur; dosya degismedikce cache'ten doner (mtime anahtarli)."""
    return _load_input_cached(path, _path_mtime(path))


# --------------------------------------------------
#  FIX #24 - NaN / Infinity sanitizasyon
# --------------------------------------------------

def _sanitize(obj):
    """JSON standardini ihlal eden float degerleri (NaN, Inf) None'a cevirir."""
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize(i) for i in obj]
    return obj


# --------------------------------------------------
#  Modeller (Pydantic) - FIX #42 strict=True
# --------------------------------------------------

class OptimizeRequest(BaseModel):
    model_config = ConfigDict(strict=True)  # FIX #42

    tarih: str = Field(..., description="Planlama tarihi (YYYY-MM-DD)")
    time_limit: int = Field(540, ge=1, description="OR-Tools zaman siniri (saniye)")

    # FIX #20 - tarih formati dogrulamasi
    @field_validator("tarih")
    @classmethod
    def validate_tarih(cls, v: str) -> str:
        try:
            datetime.date.fromisoformat(v)
        except ValueError:
            raise ValueError("tarih YYYY-MM-DD formatinda olmali (orn: 2026-01-15)")
        return v


class RentalAssignmentResponse(BaseModel):
    model_config = ConfigDict(strict=False)

    vehicle_id: str
    origin: str
    destination: str
    assigned_desi: float
    capacity_desi: float
    utilisation: float
    cost: float
    cost_type: str


class SpotAssignmentResponse(BaseModel):
    model_config = ConfigDict(strict=False)

    vehicle_type: str
    origin: str
    destination: str
    assigned_desi: float
    capacity_desi: float
    utilisation: float
    cost: float
    route_path: list[str]
    source: str


class OptimizeResponse(BaseModel):
    model_config = ConfigDict(strict=False)

    date: str
    solver_status: str
    total_rental_cost: float
    total_spot_cost: float
    total_cost: float
    fallback_count: int
    unassigned_demand: dict
    rental_assignments: list[RentalAssignmentResponse]
    spot_assignments: list[SpotAssignmentResponse]
    calisma_suresi_sn: float


# FIX #34 - Sahte kapasite/asim/ceza yerine gercek talep yogunlugu modeli
class SehirTalepYogunlugu(BaseModel):
    model_config = ConfigDict(strict=False)

    sehir_id: str
    sehir_ad: str
    giren_desi: int   # sehre gelen toplam desi (varis)
    cikan_desi: int   # sehirden cikan toplam desi (kaynak)
    toplam_desi: int  # giren + cikan


class FleetVehicle(BaseModel):
    model_config = ConfigDict(strict=False)

    arac_id: str
    tip: str
    sabit_gunluk: float
    aktif: bool
    rota: Optional[str] = None
    doluluk_pct: float = 0.0


# --------------------------------------------------
#  Yardimci Fonksiyonlar
# --------------------------------------------------

# FIX #4 - _COORDS_XLSX bagimliligi kaldirildi
# Koordinatlar artik logiai_mvp_input.json["city_coords"] anahtarindan okunur.
# Format: {"Istanbul": {"lat": 41.01, "lon": 28.97}, ...}

@lru_cache(maxsize=2)
def _load_city_coords_cached(mtime: float) -> dict[str, dict]:
    try:
        data = load_input(INPUT_JSON)
        return data.get("city_coords", {})
    except Exception as exc:
        logger.warning("Koordinat verisi okunamadi: %s", exc)
        return {}


def _load_city_coords() -> dict[str, dict]:
    """
    city_coords'u logiai_mvp_input.json'dan okur (FIX #39 - mtime anahtarli).
    JSON'da "city_coords" anahtari yoksa bos dict doner.
    """
    return _load_city_coords_cached(_input_mtime())


# FIX #1 - int(float(desi)) -> float round(2)
# FIX #21 - dict index: tarih -> list[row] ile O(1) erisim

@lru_cache(maxsize=2)
def _build_demand_index_cached(mtime: float) -> dict[str, list[dict]]:
    """
    daily_demand'i {tarih: [row, ...]} formatinda indeksler.
    FIX #39 - mtime anahtarli; girdi degisince yeniden kurulur.
    """
    data = load_input(INPUT_JSON)
    daily_demand = data.get("daily_demand", {})
    dates_sorted = sorted(daily_demand.keys())
    index: dict[str, list[dict]] = {}
    for day_idx, tarih in enumerate(dates_sorted, 1):
        rows = []
        for origin, dests in daily_demand[tarih].items():
            for dest, desi in dests.items():
                val = round(float(desi), 2)  # FIX #1
                if val > 0:
                    rows.append({
                        "tarih": tarih,
                        "gonderen_id": origin,
                        "alan_id": dest,
                        "talep_desi": str(val),  # FIX #1 - ondalik korundu
                        "gun": str(day_idx),
                    })
        index[tarih] = rows
    return index


def _build_demand_index() -> dict[str, list[dict]]:
    """FIX #39 - mtime anahtarli demand index (girdi degisince tazelenir)."""
    return _build_demand_index_cached(_input_mtime())


def load_demand() -> list[dict]:
    """Tum talep satirlarini duz liste olarak doner."""
    idx = _build_demand_index()
    result = []
    for rows in idx.values():
        result.extend(rows)
    return result


def load_distance_matrix() -> dict:
    """logiai_mvp_input.json'dan mesafe matrisini doner."""
    return load_input(INPUT_JSON).get("distance_matrix", {})


def load_travel_time_matrix() -> dict:
    """Seyahat suresi matrisini mesafeden turetir (ort. 80 km/saat)."""
    dist = load_distance_matrix()
    return {
        origin: {dest: round(km / 80.0, 2) for dest, km in dests.items()}
        for origin, dests in dist.items()
    }


def get_demand_for_date(date_str: str) -> list[dict]:
    """FIX #21 - O(1) dict lookup, O(N) tarama yok."""
    return _build_demand_index().get(date_str, [])


def _run_pipeline(data: dict, date: str, time_limit_sec: int = 540) -> dict:
    """
    Iki asamali optimizasyon boru hattini calistirir ve JSON-uyumlu dict doner.

    Asama 1 -> Greedy kiralik atama   (optimization.greedy)
    Asama 2 -> OR-Tools Open VRP      (optimization.vrp_solver)
    """
    res = run_pipeline(data, date, time_limit_sec=time_limit_sec)
    raw = result_to_dict(res)
    return _sanitize(raw)  # FIX #24


# --------------------------------------------------
#  Endpoints
# --------------------------------------------------

@app.get("/")
def root():
    return {
        "sistem": "Lojistik Optimizasyon API",
        "versiyon": "1.0.0",
        "durum": "aktif",
        "zaman": datetime.datetime.now(datetime.timezone.utc).isoformat(),  # FIX #31
    }


@app.get("/api/health")
def health():
    return {
        "status": "healthy",
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),  # FIX #31
    }


# FIX #3 - Senkron endpoint deprecated; async kullanimi oneriliyor
@app.post("/api/optimize", response_model=OptimizeResponse, deprecated=True)
def optimize(req: OptimizeRequest):
    """
    DEPRECATED - Bu endpoint 9 dakikaya kadar bloke olabilir.
    Lutfen /api/optimize/async endpoint'ini kullanin.

    Iki Asamali Optimizasyon Motoru (On Rapor Uyumlu)

    Asama 1 - Greedy Kiralik Filo Atama
    Asama 2 - OR-Tools Spot VRP
    """
    start = datetime.datetime.now(datetime.timezone.utc)  # FIX #31

    if not os.path.exists(INPUT_JSON):
        raise HTTPException(404, f"Girdi dosyasi bulunamadi: {INPUT_JSON}")

    try:
        data = load_input(INPUT_JSON)
    except DataContractError as e:
        raise HTTPException(400, f"Veri sozlesmesi hatasi: {e}")

    dates = available_dates(data)
    if req.tarih not in dates:
        raise HTTPException(
            404,
            f"{req.tarih} tarihi icin talep verisi bulunamadi. "
            f"Mevcut tarihler: {dates}"
        )

    result = _run_pipeline(data, req.tarih, time_limit_sec=req.time_limit)

    elapsed = (datetime.datetime.now(datetime.timezone.utc) - start).total_seconds()
    result["calisma_suresi_sn"] = round(elapsed, 3)

    return result


@app.get("/api/predict")
def predict(
    tarih: str = Query(..., description="Tahmin tarihi (YYYY-MM-DD)"),
    sehir: Optional[str] = Query(None, description="Sehir filtresi"),
):
    """
    LSTM Talep Tahmini (Placeholder)
    Gelismis asamada gercek LSTM modeli entegre edilecek.
    Simdilik historical data uzerinden basit istatistiksel tahmin.
    """
    demand_data = load_demand()

    from collections import defaultdict
    city_demand: dict[str, list[int]] = defaultdict(list)
    for r in demand_data:
        key = r["gonderen_id"] if sehir is None or r["gonderen_id"] == sehir else None
        if key:
            city_demand[key].append(int(float(r["talep_desi"])))

    predictions = []
    for city, demands in city_demand.items():
        avg = sum(demands) / len(demands) if demands else 0
        std = (sum((d - avg) ** 2 for d in demands) / len(demands)) ** 0.5 if demands else 0
        predictions.append({
            "sehir": city,
            "tahmin_tarih": tarih,
            "p50": round(avg, 0),
            "p10": round(max(0, avg - 1.28 * std), 0),
            "p90": round(avg + 1.28 * std, 0),
            "model": "statistical_baseline",
            "not": "LSTM modeli gelismis asamada entegre edilecektir",
        })

    return {"tarih": tarih, "tahminler": predictions}


# FIX #41 - Statik filo iskeleti (rental_routes + cost_matrix taramasi) mtime
# anahtariyla cache'lenir; her istekte yeniden taranmaz. Tarihe/job'a bagli
# doluluk_pct ise runtime'da uzerine bindirilir (job verisi cache'lenmez).
@lru_cache(maxsize=2)
def _fleet_skeleton(mtime: float) -> tuple[dict, ...]:
    data = load_input(INPUT_JSON)
    vehicles_info = data.get("vehicles_info", {})
    rows = []
    for route_key, vehicles in data.get("rental_routes", {}).items():
        origin, dest = route_key.split("_", 1)
        for v in vehicles:
            vtype = v.get("vehicle_type", "TIR")
            # FIX #64 - rental_fixed_daily_cost gercek gunluk sabit maliyeti verir.
            # cost_matrix["kiralik"] tam sefer maliyetidir, yanlis gosteriliyordu.
            vi = vehicles_info.get(vtype, {})
            if vi:
                sabit = float(vi.get("rental_fixed_daily_cost", 0))
            else:
                # Eski JSON (vehicles_info yok): cost_matrix fallback
                cost_row = data.get("cost_matrix", {}).get(origin, {}).get(dest, {}).get(vtype, {})
                sabit = float(cost_row.get("kiralik", cost_row.get("kiralık", 0)))
            rows.append({
                "arac_id": v["id"],
                "tip": vtype,
                "sabit_gunluk": sabit,
                "rota": f"{origin}->{dest}",
            })
    return tuple(rows)



@app.get("/api/fleet", response_model=list[FleetVehicle])
def fleet_status(tarih: Optional[str] = Query(None)):
    """Kiralik filo durum raporu. FIX #41 - iskelet cache'li, doluluk runtime."""
    utilisation_map: dict[str, float] = {}
    if tarih:
        job = _get_job_for_date(tarih)
        if job and job.get("status") == "COMPLETED":
            for a in job["result"].get("rental_assignments", []):
                vid = a["vehicle_id"]
                utilisation_map[vid] = min(
                    utilisation_map.get(vid, 0.0) + a["utilisation"], 1.0
                )

    result = []
    for row in _fleet_skeleton(_input_mtime()):
        vid = row["arac_id"]
        result.append(FleetVehicle(
            arac_id=vid,
            tip=row["tip"],
            sabit_gunluk=row["sabit_gunluk"],
            aktif=True,
            rota=row["rota"],
            doluluk_pct=round(utilisation_map.get(vid, 0.0) * 100, 1),
        ))
    return result


# FIX #41 - Sonuc (mtime, tarih) anahtariyla cache'lenir; her istekte yeniden
# hesaplanmaz, girdi degisince otomatik tazelenir.
@lru_cache(maxsize=256)
def _tm_density(mtime: float, target_date: str) -> list[SehirTalepYogunlugu]:
    data = load_input(INPUT_JSON)
    daily_demand = data.get("daily_demand", {})

    giren: dict[str, float] = {}
    cikan: dict[str, float] = {}
    for origin, dests in daily_demand.get(target_date, {}).items():
        for dest, desi in dests.items():
            cikan[origin] = cikan.get(origin, 0.0) + float(desi)
            giren[dest]   = giren.get(dest,   0.0) + float(desi)

    result = []
    for city in sorted(set(giren) | set(cikan)):
        g = int(giren.get(city, 0.0))
        c = int(cikan.get(city, 0.0))
        result.append(SehirTalepYogunlugu(
            sehir_id=city[:4].upper(),
            sehir_ad=city,
            giren_desi=g,
            cikan_desi=c,
            toplam_desi=g + c,
        ))
    return result


@app.get("/api/tm-status", response_model=list[SehirTalepYogunlugu])
def tm_status(tarih: Optional[str] = Query(None)):
    """
    FIX #34 - Sehir bazli talep yogunlugu (gercek daily_demand'den).
    MVP/Dataset A'da TM kapasite kisiti yoktur (sartname); bu yuzden uydurma
    kapasite/asim/ceza yerine secili tarih icin sehre giren (varis) ve
    sehirden cikan (kaynak) gercek desi gosterilir.
    FIX #41 - Sonuc cache'lenir.
    """
    # FIX #31 - UTC-aware datetime
    target_date = tarih or datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d")
    return _tm_density(_input_mtime(), target_date)


def _build_excel_zip_bytes(start_str: str, gun_sayisi: int, time_limit: int) -> bytes:
    """
    FIX #50 (E/F/G) - Iki ayri xlsx (ZIP) cekirdegi; HTTP'den bagimsiz, bytes doner.
    Hem sync /api/excel hem async worker (#77) bunu kullanir.

    Uretilen dosyalar (ZIP icinde):
      1. 1_Tahmin_Talep_Ciktisi.xlsx - preprocessing'in koke yazdigi forecast'ten kopyalanir
      2. 2_Arac_Planlama_Ciktisi.xlsx - gun_sayisi gun optimizasyon (juri formati + Maliyet Ozeti)
    """
    import io
    import zipfile
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # --- 0. Ortak stil nesneleri ---
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center")

    mvp_data = load_input(INPUT_JSON)

    # rental_routes id -> vehicle_type eslemes
    id_to_vtype: dict[str, str] = {}
    for vehicles in mvp_data.get("rental_routes", {}).values():
        for v in vehicles:
            id_to_vtype[v["id"]] = v.get("vehicle_type", "TIR")

    # --- 1. DOSYA A: 1_Tahmin_Talep_Ciktisi.xlsx ---
    # Yeni yapi: preprocessing forecast'i repo koküne 1_Tahmin_Talep_Ciktisi.xlsx
    # olarak yazar; mevcutsa dogrudan oku.
    # FIX #76 - Forecast kaynagi sira: data/processed (Docker volume-mount) ->
    # proje koku -> eski ad. Kok dizin container'a kopyalanmadigindan once
    # OUTPUT_DIR (data/processed) denenir.
    talep_candidates = [
        os.path.join(OUTPUT_DIR, "1_Tahmin_Talep_Ciktisi.xlsx"),
        os.path.join(_PROJECT_ROOT, "1_Tahmin_Talep_Ciktisi.xlsx"),
        os.path.join(OUTPUT_DIR, "Tahminlenen_Talep.xlsx"),
    ]
    talep_src = next((p for p in talep_candidates if os.path.exists(p)), None)
    if talep_src:
        wb_talep = openpyxl.load_workbook(talep_src)
        ws_talep = wb_talep.active
        for col, h in enumerate(["Tarih", "Çıkış TM", "Varış TM", "Tahmin Edilen Desi"], 1):
            c = ws_talep.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
    else:
        # FIX #76 - Forecast dosyasi yoksa daily_demand'den uret AMA yalnizca
        # planlanan gun penceresini (start_str + gun_sayisi) yaz; tum 137 gunu DOKME.
        try:
            _sd = datetime.date.fromisoformat(start_str)
        except ValueError:
            _sd = datetime.date(2026, 5, 11)
        _window = {(_sd + datetime.timedelta(days=i)).isoformat() for i in range(gun_sayisi)}
        wb_talep = Workbook()
        ws_talep = wb_talep.active
        ws_talep.title = "Tahminlenen Talep"
        for col, h in enumerate(["Tarih", "Çıkış TM", "Varış TM", "Tahmin Edilen Desi"], 1):
            c = ws_talep.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
        row_idx = 2
        for row in load_demand():
            if row["tarih"] not in _window:
                continue
            ws_talep.cell(row=row_idx, column=1, value=row["tarih"])
            ws_talep.cell(row=row_idx, column=2, value=row["gonderen_id"])
            ws_talep.cell(row=row_idx, column=3, value=row["alan_id"])
            ws_talep.cell(row=row_idx, column=4, value=float(row["talep_desi"]))
            row_idx += 1

    # --- 2. DOSYA B: Arac_Planlama.xlsx ---
    try:
        start_date = datetime.date.fromisoformat(start_str)
    except ValueError:
        raise HTTPException(400, f"Gecersiz tarih formati: {start_str}. YYYY-MM-DD olmali.")

    # Planlama tarihlerini olustur
    all_dates_in_data = sorted(available_dates(mvp_data))
    plan_dates = []
    for i in range(gun_sayisi):
        d = (start_date + datetime.timedelta(days=i)).isoformat()
        if d in all_dates_in_data:
            plan_dates.append(d)

    if not plan_dates:
        raise HTTPException(404, f"{start_str} tarihinden itibaren {gun_sayisi} gun icinde veri bulunamadi.")

    wb_plan = Workbook()
    ws_plan = wb_plan.active
    ws_plan.title = "Arac Planlama"

    plan_headers = ["Tarih", "Araç Tipi", "Çıkış TM", "Varış TM", "Atanan Desi", "Maliyet", "Tür"]
    for col, h in enumerate(plan_headers, 1):
        c = ws_plan.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    ws_ozet = wb_plan.create_sheet("Maliyet Ozeti")
    ozet_headers = ["Tarih", "Kiralik Toplam (TL)", "Spot Toplam (TL)", "Genel Toplam (TL)"]
    for col, h in enumerate(ozet_headers, 1):
        c = ws_ozet.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    plan_row  = 2
    ozet_row  = 2
    grand_rental = 0.0
    grand_spot   = 0.0

    for gun_tarih in plan_dates:
        # Optimizasyon sonucunu al: once tamamlanmis async job, yoksa sync calistir
        result = None
        job = _get_job_for_date(gun_tarih)
        if job and job.get("status") == "COMPLETED":
            cached = job.get("result") or {}
            # FIX - bos/bayat cache'i kullanma; atama yoksa yeniden hesapla
            if cached.get("rental_assignments") or cached.get("spot_assignments"):
                result = cached
        if result is None:
            try:
                result = _run_pipeline(mvp_data, gun_tarih, time_limit_sec=time_limit)
            except Exception as exc:
                logger.warning("Excel: %s icin optimizasyon basarisiz: %s", gun_tarih, exc)
                result = {"rental_assignments": [], "spot_assignments": []}

        rental_assignments = result.get("rental_assignments", [])
        spot_assignments   = result.get("spot_assignments", [])
        day_rental = 0.0
        day_spot   = 0.0

        for a in rental_assignments:
            vtype = id_to_vtype.get(a.get("vehicle_id", ""), "TIR")
            cost  = float(a.get("cost", 0.0))
            day_rental += cost
            ws_plan.cell(row=plan_row, column=1, value=gun_tarih)
            ws_plan.cell(row=plan_row, column=2, value=f"Kiralık {vtype}")
            ws_plan.cell(row=plan_row, column=3, value=a.get("origin", ""))
            ws_plan.cell(row=plan_row, column=4, value=a.get("destination", ""))
            ws_plan.cell(row=plan_row, column=5, value=round(float(a.get("assigned_desi", 0.0)), 2))
            ws_plan.cell(row=plan_row, column=6, value=round(cost, 2))
            ws_plan.cell(row=plan_row, column=7, value="Kiralık")
            plan_row += 1

        for a in spot_assignments:
            cost = float(a.get("cost", 0.0))
            day_spot += cost
            ws_plan.cell(row=plan_row, column=1, value=gun_tarih)
            ws_plan.cell(row=plan_row, column=2, value=f"Spot {a.get('vehicle_type', '')}")
            ws_plan.cell(row=plan_row, column=3, value=a.get("origin", ""))
            ws_plan.cell(row=plan_row, column=4, value=a.get("destination", ""))
            ws_plan.cell(row=plan_row, column=5, value=round(float(a.get("assigned_desi", 0.0)), 2))
            ws_plan.cell(row=plan_row, column=6, value=round(cost, 2))
            ws_plan.cell(row=plan_row, column=7, value="Spot")
            plan_row += 1

        ws_ozet.cell(row=ozet_row, column=1, value=gun_tarih)
        ws_ozet.cell(row=ozet_row, column=2, value=round(day_rental, 2))
        ws_ozet.cell(row=ozet_row, column=3, value=round(day_spot, 2))
        ws_ozet.cell(row=ozet_row, column=4, value=round(day_rental + day_spot, 2))
        ozet_row += 1

        grand_rental += day_rental
        grand_spot   += day_spot

    # Toplam satiri
    ws_ozet.cell(row=ozet_row, column=1, value="GENEL TOPLAM")
    ws_ozet.cell(row=ozet_row, column=2, value=round(grand_rental, 2))
    ws_ozet.cell(row=ozet_row, column=3, value=round(grand_spot, 2))
    ws_ozet.cell(row=ozet_row, column=4, value=round(grand_rental + grand_spot, 2))
    for col in range(1, 5):
        ws_ozet.cell(row=ozet_row, column=col).font = Font(bold=True)

    # --- 3. ZIP icinde sun ---
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        talep_buf = io.BytesIO()
        wb_talep.save(talep_buf)
        zf.writestr("1_Tahmin_Talep_Ciktisi.xlsx", talep_buf.getvalue())

        plan_buf = io.BytesIO()
        wb_plan.save(plan_buf)
        zf.writestr("2_Arac_Planlama_Ciktisi.xlsx", plan_buf.getvalue())

    return zip_buf.getvalue()


@app.get("/api/excel")
def generate_excel(
    tarih: Optional[str] = Query(None, description="Baslangic tarihi (YYYY-MM-DD); verilmezse 2026-05-11"),
    time_limit: int = Query(120, ge=1, description="Her gun icin sync optimizasyon zaman siniri (sn)"),
    gun_sayisi: int = Query(7, ge=1, le=30, description="Kac gunluk planlama uretilecek"),
):
    """
    Senkron Excel uretimi. Buyuk gun_sayisi (orn. 7) optimizasyon kostugundan
    dakikalarca surebilir; dashboard 30sn timeout'una takilirsa /api/excel/async kullanin (#77).
    """
    import io as _io
    from fastapi.responses import StreamingResponse
    start_str = tarih or "2026-05-11"
    zip_bytes = _build_excel_zip_bytes(start_str, gun_sayisi, time_limit)
    return StreamingResponse(
        _io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="logiai_cikti_{start_str}.zip"'},
    )


@app.post("/api/excel/async")
def excel_async(
    background_tasks: BackgroundTasks,
    tarih: Optional[str] = Query(None, description="Baslangic tarihi (YYYY-MM-DD); verilmezse 2026-05-11"),
    time_limit: int = Query(120, ge=1, description="Her gun icin optimizasyon zaman siniri (sn)"),
    gun_sayisi: int = Query(7, ge=1, le=30, description="Kac gunluk planlama uretilecek"),
):
    """
    FIX #77 - Excel uretimini arka planda calistirir; istek aninda bloke olmaz
    (7 gun x time_limit = dakikalar surebilir, dashboard 30sn timeout'una takilmaz).

    Hemen job_id doner. Istemci /api/jobs/{job_id} ile durumu izler;
    COMPLETED olunca /api/excel/result/{job_id} ile ZIP'i indirir.
    """
    start_str = tarih or "2026-05-11"
    try:
        datetime.date.fromisoformat(start_str)
    except ValueError:
        raise HTTPException(400, f"Gecersiz tarih formati: {start_str}. YYYY-MM-DD olmali.")

    if not _semaphore.acquire(blocking=False):
        raise HTTPException(429, f"Maksimum eszamanli is sayisina ({_MAX_CONCURRENT_JOBS}) ulasildi.")

    job_id = create_job()

    def _worker():
        set_running(job_id)
        try:
            zip_bytes = _build_excel_zip_bytes(start_str, gun_sayisi, time_limit)
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            path = os.path.join(OUTPUT_DIR, f"logiai_cikti_{start_str}.zip")
            with open(path, "wb") as fh:
                fh.write(zip_bytes)
            set_completed(job_id, {"file": path, "date": start_str})
        except Exception as exc:
            logger.exception("Excel worker hatasi job_id=%s", job_id)
            try:
                set_failed(job_id, str(exc))
            except Exception as inner:
                logger.error("set_failed basarisiz job_id=%s: %s", job_id, inner)
        finally:
            _semaphore.release()

    background_tasks.add_task(_worker)
    return {"job_id": job_id, "status": "PENDING"}


@app.get("/api/excel/result/{job_id}")
def excel_result(job_id: str):
    """FIX #77 - Tamamlanmis Excel job'unun ZIP dosyasini indirir."""
    from fastapi.responses import FileResponse
    job = _get_job(job_id)
    if job is None:
        raise HTTPException(404, "Job bulunamadi veya suresi doldu (TTL: 1 saat).")
    status = job.get("status")
    if status == "FAILED":
        raise HTTPException(500, f"Excel uretimi basarisiz: {job.get('error', '?')}")
    if status != "COMPLETED":
        raise HTTPException(409, f"Excel henuz hazir degil (durum: {status}).")
    path = (job.get("result") or {}).get("file")
    if not path or not os.path.exists(path):
        raise HTTPException(404, "Cikti dosyasi bulunamadi.")
    return FileResponse(path, media_type="application/zip", filename=os.path.basename(path))



@app.get("/api/cities")
def list_cities():
    """
    Sehir listesi ve koordinatlari (logiai_mvp_input.json'dan).
    FIX #34 - Uydurma tm_var/tm_kapasite/tir_yanasma sabitleri kaldirildi;
    MVP'de TM kapasite kisiti yoktur, bu alanlar veriye dayanmiyordu.
    """
    # FIX #4 - Excel bagimliligi kaldirildi
    data = load_input(INPUT_JSON)
    coords_map = _load_city_coords()
    cities = []
    for city_name in sorted(data["distance_matrix"].keys()):
        coords = coords_map.get(city_name, {"lat": 39.0, "lon": 35.0})
        cities.append({
            "id":   city_name,
            "ad":   city_name,
            "lat":  coords["lat"],
            "lon":  coords["lon"],
        })
    return {"sehirler": cities}


@app.get("/api/dates")
def list_dates():
    """
    Talep verisinde mevcut planlama tarihleri (sirali).
    Dashboard tarih secici bu listeden beslenir; boylece veride olmayan
    bir tarih secilip 404 alinmasi engellenir.
    """
    data = load_input(INPUT_JSON)
    return {"tarihler": sorted(available_dates(data))}


@app.get("/api/vehicles")
def list_vehicles():
    """
    FIX #32 - Arac tipi bilgileri vehicles_info'dan okunuyor (gercek kapasite+maliyet).
    vehicle_types varsa o kullanilir; yoksa vehicles_info yapilandirilir;
    o da yoksa cost_matrix'ten benzersiz tipler + statik fallback devreye girer.
    """
    data = load_input(INPUT_JSON)

    # Oncelik 1: vehicle_types anahtari
    if "vehicle_types" in data:
        return {"arac_tipleri": data["vehicle_types"]}

    # Oncelik 2: vehicles_info (FIX #32 - gercek kapasite ve maliyet)
    vehicles_info = data.get("vehicles_info", {})
    if vehicles_info:
        # tir_yanasma JSON'daki transfer_centers'ten anlasilabilir; JSON'da
        # vehicle_types yoksa kodu koda donustur
        _tir_required = {"TIR": True, "KAM": False, "HAF": False, "KMT": False}
        arac_tipleri = []
        for code, info in vehicles_info.items():
            arac_tipleri.append({
                "id": code,
                "ad": info.get("name", code),
                "kapasite_desi": info.get("capacity_desi", 0),
                "sabit_maliyet": info.get("rental_fixed_daily_cost", 0.0),
                "km_basi_maliyet": info.get("rental_cost_per_km", 0.0),
                "spot_sabit_maliyet": info.get("spot_fixed_daily_cost", 0.0),
                "spot_km_basi_maliyet": info.get("spot_cost_per_km", 0.0),
                "tir_yanasma_gerekli": _tir_required.get(code, False),
            })
        return {"arac_tipleri": arac_tipleri}

    # Fallback: cost_matrix'ten benzersiz arac tiplerini topla + statik tablo
    cost_matrix = data.get("cost_matrix", {})
    seen: set[str] = set()
    for origin_data in cost_matrix.values():
        for dest_data in origin_data.values():
            seen.update(dest_data.keys())

    _static_caps = {
        "TIR": {"id": "TIR", "ad": "Tir", "kapasite_desi": 22400, "sabit_maliyet": 7000.0, "km_basi_maliyet": 13.0, "tir_yanasma_gerekli": True},
        "KAM": {"id": "KAM", "ad": "Kamyon", "kapasite_desi": 12000, "sabit_maliyet": 5000.0, "km_basi_maliyet": 10.0, "tir_yanasma_gerekli": False},
        "HAF": {"id": "HAF", "ad": "Hafif Kamyon", "kapasite_desi": 7200, "sabit_maliyet": 5000.0, "km_basi_maliyet": 10.0, "tir_yanasma_gerekli": False},
        "KMT": {"id": "KMT", "ad": "Kamyonet", "kapasite_desi": 5600, "sabit_maliyet": 3750.0, "km_basi_maliyet": 6.0, "tir_yanasma_gerekli": False},
    }
    arac_tipleri = [_static_caps[t] for t in seen if t in _static_caps]
    if not arac_tipleri:
        arac_tipleri = list(_static_caps.values())

    return {"arac_tipleri": arac_tipleri}


# FIX #33 - pagination eklendi
@app.get("/api/demand")
def get_demand(
    tarih: Optional[str] = Query(None),
    sehir: Optional[str] = Query(None),
    limit: int = Query(500, ge=1, le=5000, description="Sayfa basi kayit"),
    offset: int = Query(0, ge=0, description="Baslangic offset"),
):
    """Talep verisi sorgulama - sayfali (limit/offset)"""
    # FIX #21 - index'li erisim
    if tarih:
        demand_data = list(_build_demand_index().get(tarih, []))
    else:
        demand_data = load_demand()

    if sehir:
        demand_data = [r for r in demand_data if r["gonderen_id"] == sehir or r["alan_id"] == sehir]

    total = len(demand_data)
    page = demand_data[offset: offset + limit]
    return {"toplam_kayit": total, "limit": limit, "offset": offset, "talepler": page}


# --------------------------------------------------
#  Async Optimizasyon - Redis Polling
# --------------------------------------------------

class AsyncJobResponse(BaseModel):
    model_config = ConfigDict(strict=False)

    job_id: str
    status: str


# FIX #44 - BackgroundTasks ile worker
# FIX #10 - Semaphore: maksimum eszamanli is siniri
@app.post("/api/optimize/async", response_model=AsyncJobResponse)
def optimize_async(req: OptimizeRequest, background_tasks: BackgroundTasks):
    """
    Optimizasyonu arka planda baslatir; arayuz kilitlenmez.

    Hemen job_id doner. Istemci /api/jobs/{job_id} adresini
    periyodik olarak sorgulayarak durumu takip eder.

    Is durumlari: PENDING -> RUNNING -> COMPLETED | FAILED

    FIX #39 - NOT: uvicorn --workers > 1 ile calistirildiginda her worker
    kendi izole havuzuna sahiptir. Gercek multi-worker ortami icin
    Celery veya ARQ kullanilmasi onerilir.
    """
    if not os.path.exists(INPUT_JSON):
        raise HTTPException(404, f"Girdi dosyasi bulunamadi: {INPUT_JSON}")

    try:
        data = load_input(INPUT_JSON)
    except DataContractError as e:
        raise HTTPException(400, f"Veri sozlesmesi hatasi: {e}")

    dates = available_dates(data)
    if req.tarih not in dates:
        raise HTTPException(
            404,
            f"{req.tarih} tarihi icin talep verisi bulunamadi. "
            f"Mevcut tarihler: {dates}",
        )

    # FIX #10 - Semaphore kapida - doluysa 429
    if not _semaphore.acquire(blocking=False):
        raise HTTPException(
            status_code=429,
            detail=f"Maksimum eszamanli is sayisina ({_MAX_CONCURRENT_JOBS}) ulasildi. Lutfen bekleyin.",
        )

    job_id = create_job()

    # FIX #44 - BackgroundTasks (FastAPI native)
    def _worker():
        set_running(job_id)
        try:
            start = datetime.datetime.now(datetime.timezone.utc)  # FIX #31
            result = _run_pipeline(data, req.tarih, time_limit_sec=req.time_limit)
            elapsed = (datetime.datetime.now(datetime.timezone.utc) - start).total_seconds()
            result["calisma_suresi_sn"] = round(elapsed, 3)
            set_completed(job_id, result)
        except Exception as exc:
            logger.exception("Worker hatasi job_id=%s", job_id)
            # FIX #6 - set_failed try/except ile guard
            try:
                set_failed(job_id, str(exc))
            except Exception as inner_exc:
                logger.error("set_failed basarisiz job_id=%s: %s", job_id, inner_exc)
        finally:
            _semaphore.release()

    background_tasks.add_task(_worker)
    return {"job_id": job_id, "status": "PENDING"}


@app.get("/api/jobs/{job_id}")
def get_job_status(job_id: str):
    """
    Is durumunu doner.

    Donen alan "status": PENDING | RUNNING | COMPLETED | FAILED
    COMPLETED ise "result" alani OptimizeResponse semasiyla aynidir.
    FAILED ise "error" alani hata mesajini icerir.
    """
    job = _get_job(job_id)
    if job is None:
        raise HTTPException(404, "Job bulunamadi veya suresi doldu (TTL: 1 saat).")
    return job


if __name__ == "__main__":
    import uvicorn
    # FIX #39 uyarisi: --workers > 1 icin Celery/ARQ kullanin
    uvicorn.run(app, host="0.0.0.0", port=8000, workers=1)
